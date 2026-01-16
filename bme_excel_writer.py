"""
Bosch Sensortec BME Excel Writer
Creates formatted Excel workbooks from BME sensor data
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from typing import Dict, Any, Optional
from pathlib import Path


class BMEExcelWriter:
    """Writer for creating Excel workbooks from BME data"""
    
    # Color scheme
    HEADER_COLOR = "366092"
    SENSOR_COLORS = [
        "E6F2FF",  # Light blue - Sensor 0
        "FFE6F0",  # Light pink - Sensor 1
        "E6FFE6",  # Light green - Sensor 2
        "FFF0E6",  # Light orange - Sensor 3
        "F0E6FF",  # Light purple - Sensor 4
        "FFFFE6",  # Light yellow - Sensor 5
        "E6FFFF",  # Light cyan - Sensor 6
        "FFE6E6",  # Light red - Sensor 7
    ]
    
    def __init__(self, output_path: str):
        """
        Initialize Excel writer
        
        Args:
            output_path: Path for output Excel file
        """
        self.output_path = output_path
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def add_configuration_sheet(self, 
                               config_df: pd.DataFrame,
                               heater_df: pd.DataFrame,
                               duty_df: pd.DataFrame,
                               sensor_df: pd.DataFrame,
                               board_type: str,
                               file_name: str):
        """
        Add configuration sheet to workbook
        
        Args:
            config_df: Configuration DataFrame
            heater_df: Heater profiles DataFrame
            duty_df: Duty cycle profiles DataFrame
            sensor_df: Sensor configurations DataFrame
            board_type: Board type string
            file_name: Source file name
        """
        ws = self.workbook.create_sheet("Configuration")
        
        current_row = 1
        
        # Add title
        ws.cell(row=current_row, column=1, value="Bosch Sensortec BME Configuration")
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Source File: {file_name}")
        current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Board Type: {board_type}")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 2
        
        # Add configuration details
        if not config_df.empty:
            ws.cell(row=current_row, column=1, value="Configuration Details")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            current_row = self._write_dataframe_to_sheet(ws, config_df, current_row)
            current_row += 2
        
        # Add heater profiles
        if not heater_df.empty:
            ws.cell(row=current_row, column=1, value="Heater Profiles")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            current_row = self._write_dataframe_to_sheet(ws, heater_df, current_row)
            current_row += 2
        
        # Add duty cycle profiles
        if not duty_df.empty:
            ws.cell(row=current_row, column=1, value="Duty Cycle Profiles")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            current_row = self._write_dataframe_to_sheet(ws, duty_df, current_row)
            current_row += 2
        
        # Add sensor configurations
        if not sensor_df.empty:
            ws.cell(row=current_row, column=1, value="Sensor Configurations")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            current_row = self._write_dataframe_to_sheet(ws, sensor_df, current_row)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def add_labels_sheet(self, labels_df: pd.DataFrame):
        """
        Add labels sheet to workbook
        
        Args:
            labels_df: Labels DataFrame
        """
        if labels_df.empty:
            return
        
        ws = self.workbook.create_sheet("Labels")
        
        current_row = 1
        
        # Add title
        ws.cell(row=current_row, column=1, value="Label Definitions")
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 2
        
        # Write DataFrame
        self._write_dataframe_to_sheet(ws, labels_df, current_row, apply_table=True)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def add_sensor_data_sheet(self, data_df: pd.DataFrame, column_info: list):
        """
        Add sensor data sheet to workbook with formatting
        
        Args:
            data_df: Sensor data DataFrame
            column_info: Column definitions with units
        """
        ws = self.workbook.create_sheet("Sensor Data")
        
        current_row = 1
        
        # Add title
        ws.cell(row=current_row, column=1, value="Sensor Data")
        ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        current_row += 2
        
        if data_df.empty:
            ws.cell(row=current_row, column=1, value="No data available")
            return
        
        # Write headers with units as comments
        header_row = current_row
        for col_idx, col_name in enumerate(data_df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.HEADER_COLOR, 
                                   end_color=self.HEADER_COLOR, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add unit as comment if available
            for col_def in column_info:
                if col_def.get('name') == col_name:
                    unit = col_def.get('unit', '')
                    if unit:
                        cell.comment = Comment(f"Unit: {unit}", "BME Converter")
                    break
        
        current_row += 1
        
        # Write data rows
        for row_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=False), start=current_row):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color rows by sensor index
                if 'Sensor Index' in data_df.columns:
                    sensor_idx = data_df.iloc[row_idx - current_row]['Sensor Index']
                    if isinstance(sensor_idx, int) and 0 <= sensor_idx < len(self.SENSOR_COLORS):
                        cell.fill = PatternFill(start_color=self.SENSOR_COLORS[sensor_idx],
                                              end_color=self.SENSOR_COLORS[sensor_idx],
                                              fill_type="solid")
        
        # Add auto-filter
        ws.auto_filter.ref = f"A{header_row}:{self._get_column_letter(len(data_df.columns))}{current_row + len(data_df) - 1}"
        
        # Freeze top rows
        ws.freeze_panes = ws[f"A{current_row}"]
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def save(self):
        """Save workbook to file"""
        self.workbook.save(self.output_path)
        return self.output_path
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, start_row: int, 
                                  apply_table: bool = False) -> int:
        """
        Write DataFrame to worksheet
        
        Args:
            ws: Worksheet object
            df: DataFrame to write
            start_row: Starting row number
            apply_table: Whether to apply table formatting
            
        Returns:
            Next available row number
        """
        header_row = start_row
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                   end_color=self.HEADER_COLOR,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 
                                      start=start_row + 1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply table formatting if requested
        if apply_table and not df.empty:
            tab = Table(displayName=f"Table{start_row}",
                       ref=f"A{header_row}:{self._get_column_letter(len(df.columns))}{start_row + len(df)}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)
        
        return start_row + len(df) + 1
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _get_column_letter(col_num: int) -> str:
        """Convert column number to letter (1 -> A, 27 -> AA, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + 65) + result
            col_num //= 26
        return result
